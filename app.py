
import streamlit as st
import numpy as np
import pandas as pd
from io import BytesIO
from fpdf import FPDF
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import plotly.graph_objects as go
import math

st.set_page_config(page_title="PV Halter Wirtschaftlichkeit (Plotly)", layout="wide")
st.title("📊 Wirtschaftlichkeitsanalyse – PV-Modulhalter (Plotly, PDF-Fix)")

# ---------- Helpers ----------
def ensure_latin1(text):
    if text is None:
        return ""
    if not isinstance(text, str):
        text = str(text)
    # Replace common Unicode chars that break fpdf (latin-1) – esp. Euro sign
    text = text.replace("€", " EUR ")
    # Fallback: strip any remaining non-latin1 characters
    return text.encode("latin-1", "ignore").decode("latin-1")

def safe_num(x, ndigits=2):
    if x is None or (isinstance(x, float) and (math.isnan(x) or math.isinf(x))):
        return ""
    try:
        return f"{float(x):.{ndigits}f}"
    except Exception:
        return ensure_latin1(x)

# ---------- Inputs ----------
st.markdown("### 1) Detailkosten & Verkaufspreis-Konfiguration")

def kosten_eingabe(titel, prefix):
    with st.expander(titel, expanded=True):
        col1, col2 = st.columns(2)
        with col1:
            material = st.number_input(f"{prefix} Materialkosten (€)", value=10.0)
            lohn = st.number_input(f"{prefix} Lohnkosten (€)", value=5.0)
            energie = st.number_input(f"{prefix} Energiekosten (€)", value=2.0)
        with col2:
            lager = st.number_input(f"{prefix} Lagerhaltungskosten (€)", value=2.0)
            transport = st.number_input(f"{prefix} Transportkosten (€)", value=3.0)
            hilfs = st.number_input(f"{prefix} Hilfsstoffkosten (€)", value=1.0)
    return material + lohn + energie + lager + transport + hilfs

st.subheader("Variable Kosten je Stück")
var_beton = kosten_eingabe("Betonfuß – Variable Kosten", "Beton")
var_recyc = kosten_eingabe("Recyclingfuß – Variable Kosten", "Recycling")

def vk_kalkulation(prefix, var_kosten):
    with st.expander(f"{prefix} Verkaufspreis-Kalkulation", expanded=True):
        aufschlag_prozent = st.number_input(f"{prefix} Aufschlag (%)", value=25.0)
        vertrieb = st.number_input(f"{prefix} Vertriebskosten (€)", value=2.0)
        skonto = st.number_input(f"{prefix} Skonto (€)", value=1.0)
        marketing = st.number_input(f"{prefix} Marketingkosten (€)", value=2.0)
        vk = var_kosten * (1 + aufschlag_prozent / 100) + vertrieb + skonto + marketing
    return vk

st.subheader("Verkaufspreise je Stück")
vk_beton = vk_kalkulation("Beton", var_beton)
vk_recyc = vk_kalkulation("Recycling", var_recyc)

st.markdown("### 2) Fixkosten & Stückzahlbereich")
col1, col2 = st.columns(2)
with col1:
    fix_beton = st.number_input("Fixkosten Beton (€)", value=23000.0)
with col2:
    fix_recyc = st.number_input("Fixkosten Recycling (€)", value=26000.0)

max_stück = st.slider("Maximale Stückzahl", 100, 5000, 2000, step=100)
x = np.arange(1, max_stück + 1)

# ---------- Calculation ----------
erlös_beton = x * vk_beton
erlös_recyc = x * vk_recyc
aufwand_beton = fix_beton + x * var_beton
aufwand_recyc = fix_recyc + x * var_recyc
gewinn_beton = erlös_beton - aufwand_beton
gewinn_recyc = erlös_recyc - aufwand_recyc
roi_beton = np.divide(gewinn_beton, aufwand_beton, out=np.full_like(gewinn_beton, np.nan, dtype=float), where=aufwand_beton!=0)
roi_recyc = np.divide(gewinn_recyc, aufwand_recyc, out=np.full_like(gewinn_recyc, np.nan, dtype=float), where=aufwand_recyc!=0)

def find_zero_crossing_x(x, y):
    sign = np.sign(y)
    idx = np.where(np.diff(sign) != 0)[0]
    if len(idx) == 0:
        return None
    i = idx[0]
    x0, x1 = x[i], x[i+1]
    y0, y1 = y[i], y[i+1]
    if (y1 - y0) == 0:
        return float(x0)
    return float(x0 - y0 * (x1 - x0) / (y1 - y0))

def find_intersection_x(x, y1, y2):
    return find_zero_crossing_x(x, y1 - y2)

be_beton_x = find_zero_crossing_x(x, gewinn_beton)
be_recyc_x = find_zero_crossing_x(x, gewinn_recyc)
gain_intersect_x = find_intersection_x(x, gewinn_beton, gewinn_recyc)
roi_beton_zero_x = find_zero_crossing_x(x, roi_beton)
roi_recyc_zero_x = find_zero_crossing_x(x, roi_recyc)

def interp(xq, x, y):
    return float(np.interp(xq, x, y)) if xq is not None else None

# ---------- Charts (Plotly) ----------
st.markdown("### 3) Ergebnisse & Diagramme")
st.subheader("Break-Even & Gewinnvergleich (interaktiv)")
import plotly.graph_objects as go
fig_gain = go.Figure()
fig_gain.add_trace(go.Scatter(x=x, y=gewinn_beton, mode="lines", name="Gewinn Beton",
                              hovertemplate="Stückzahl: %{x}<br>Gewinn: %{y:.2f} EUR<extra></extra>"))
fig_gain.add_trace(go.Scatter(x=x, y=gewinn_recyc, mode="lines", name="Gewinn Recycling",
                              hovertemplate="Stückzahl: %{x}<br>Gewinn: %{y:.2f} EUR<extra></extra>"))
fig_gain.add_hline(y=0, line_dash="dash", opacity=0.6)

points = []
if be_beton_x is not None:
    points.append(("Break-Even Beton", be_beton_x, interp(be_beton_x, x, gewinn_beton)))
if be_recyc_x is not None:
    points.append(("Break-Even Recycling", be_recyc_x, interp(be_recyc_x, x, gewinn_recyc)))
if gain_intersect_x is not None:
    points.append(("Schnitt Gewinnkurven", gain_intersect_x, interp(gain_intersect_x, x, gewinn_beton)))

for label, px, py in points:
    fig_gain.add_trace(go.Scatter(
        x=[px], y=[py], mode="markers+text", name=label,
        text=[f"{label}<br>Stk≈{px:.0f}, Wert≈{py:.0f} EUR"],
        textposition="top center",
        hovertemplate=f"{label}<br>Stückzahl: %{{x:.2f}}<br>Wert: %{{y:.2f}} EUR<extra></extra>"
    ))

fig_gain.update_layout(
    xaxis_title="Stückzahl",
    yaxis_title="Gewinn (EUR)",
    hovermode="x unified",
    margin=dict(l=40, r=20, t=40, b=40)
)
st.plotly_chart(fig_gain, use_container_width=True)

st.subheader("ROI Verlauf (interaktiv)")
fig_roi = go.Figure()
fig_roi.add_trace(go.Scatter(x=x, y=roi_beton, mode="lines", name="ROI Beton",
                             hovertemplate="Stückzahl: %{x}<br>ROI: %{y:.4f}<extra></extra>"))
fig_roi.add_trace(go.Scatter(x=x, y=roi_recyc, mode="lines", name="ROI Recycling",
                             hovertemplate="Stückzahl: %{x}<br>ROI: %{y:.4f}<extra></extra>"))
fig_roi.add_hline(y=0, line_dash="dash", opacity=0.6)

roi_points = []
if roi_beton_zero_x is not None:
    roi_points.append(("ROI=0 Beton", roi_beton_zero_x, interp(roi_beton_zero_x, x, roi_beton)))
if roi_recyc_zero_x is not None:
    roi_points.append(("ROI=0 Recycling", roi_recyc_zero_x, interp(roi_recyc_zero_x, x, roi_recyc)))

for label, px, py in roi_points:
    fig_roi.add_trace(go.Scatter(
        x=[px], y=[py], mode="markers+text", name=label,
        text=[f"{label}<br>Stk≈{px:.0f}"],
        textposition="top center",
        hovertemplate=f"{label}<br>Stückzahl: %{{x:.2f}}<extra></extra>"
    ))

fig_roi.update_layout(
    xaxis_title="Stückzahl",
    yaxis_title="ROI",
    hovermode="x unified",
    margin=dict(l=40, r=20, t=40, b=40)
)
st.plotly_chart(fig_roi, use_container_width=True)

# ---------- Data Table ----------
df = pd.DataFrame({
    "Stückzahl": x,
    "Gewinn Beton (EUR)": gewinn_beton,
    "Gewinn Recycling (EUR)": gewinn_recyc,
    "ROI Beton": roi_beton,
    "ROI Recycling": roi_recyc
})
if st.checkbox("📋 Tabelle anzeigen"):
    st.dataframe(df)

# ---------- Export ----------
def to_excel(df):
    wb = Workbook()
    ws = wb.active
    ws.title = "Wirtschaftlichkeit"
    # Ensure column names ASCII-friendly for Excel too
    cols = [c.replace("€", "EUR") for c in df.columns]
    df_x = df.copy()
    df_x.columns = cols
    for r in dataframe_to_rows(df_x, index=False, header=True):
        ws.append(r)
    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio

def to_pdf(df):
    pdf = FPDF(orientation="P", unit="mm", format="A4")
    pdf.add_page()
    pdf.set_auto_page_break(auto=True, margin=12)
    pdf.set_font("Arial", size=11)
    pdf.cell(0, 8, ensure_latin1("Wirtschaftlichkeitsanalyse PV-Modulhalter"), ln=1, align="C")
    pdf.ln(2)
    # Prepare headers & rows (latin-1 safe)
    col_names = [ensure_latin1(c.replace("€", "EUR")) for c in df.columns]
    ncols = len(col_names)
    page_width = 210 - 2*10  # A4 width minus margins
    col_width = page_width / ncols
    # Header
    pdf.set_font("Arial", "B", 9)
    for col in col_names:
        pdf.cell(col_width, 8, col, border=1, align="C")
    pdf.ln()
    pdf.set_font("Arial", size=8)
    max_rows = min(len(df), 80)
    for _, row in df.head(max_rows).iterrows():
        for item in row:
            if isinstance(item, (int, float, np.floating)):
                cell_text = safe_num(item)
            else:
                cell_text = ensure_latin1(item)
            pdf.cell(col_width, 6, cell_text, border=1)
        pdf.ln()
    bio = BytesIO()
    pdf.output(bio)
    bio.seek(0)
    return bio

st.markdown("### 4) 📤 Export")
col3, col4 = st.columns(2)
with col3:
    excel_bytes = to_excel(df)
    st.download_button("📥 Excel herunterladen", data=excel_bytes, file_name="wirtschaftlichkeit_pvhalter.xlsx")

with col4:
    pdf_bytes = to_pdf(df)
    st.download_button("📥 PDF herunterladen", data=pdf_bytes, file_name="wirtschaftlichkeit_pvhalter.pdf")
